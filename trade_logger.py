"""
Append-only trade log with CSV and Excel output.

output/trade_log.csv         — one row per closed leg
output/daily_summary.csv     — rebuilt from trade_log on every write
output/APEX_Trade_Log.xlsx   — 3-sheet workbook, rebuilt on every write
"""
from __future__ import annotations

import csv
import threading
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional

from config.logger import get_logger

log = get_logger(__name__)

_OUTPUT_DIR = Path("output")
_TRADE_LOG  = _OUTPUT_DIR / "trade_log.csv"
_DAILY_LOG  = _OUTPUT_DIR / "daily_summary.csv"
_EXCEL_LOG  = _OUTPUT_DIR / "APEX_Trade_Log.xlsx"

_TRADE_COLS: list[str] = [
    "date", "symbol", "direction", "entry_time", "entry_price",
    "exit_time", "exit_price", "quantity", "gross_pnl",
    "strategy_name", "confluence_score", "stop_loss", "take_profit",
    "exit_reason", "confluences_used", "bias_score", "kill_zone",
]

_DAILY_COLS: list[str] = [
    "date", "total_trades", "winners", "losers", "gross_pnl",
    "win_rate", "best_trade", "worst_trade",
]


class TradeLogger:
    """Thread-safe, append-only trade logger with CSV and Excel output."""

    def __init__(self) -> None:
        self._lock = threading.Lock()
        _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        self._ensure_header(_TRADE_LOG, _TRADE_COLS)
        self._ensure_header(_DAILY_LOG, _DAILY_COLS)
        self._verify_excel()

    @staticmethod
    def _ensure_header(path: Path, cols: list[str]) -> None:
        if not path.exists():
            with open(path, "w", newline="") as fh:
                csv.writer(fh).writerow(cols)
            log.info("Created %s", path)

    def _verify_excel(self) -> None:
        """On startup: confirm openpyxl is available and the Excel file is writable."""
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            print(
                "WARNING: openpyxl not installed — Excel output disabled. "
                "Run: pip install openpyxl",
                flush=True,
            )
            log.warning("openpyxl not installed; Excel export disabled. Run: pip install openpyxl")
            return

        # Attempt a no-op write to confirm the file isn't locked
        try:
            if _EXCEL_LOG.exists():
                # Test write-access without modifying content
                with open(_EXCEL_LOG, "a"):
                    pass
            print(f"  Trade logger ready : {_EXCEL_LOG.resolve()}", flush=True)
            log.info("Trade logger ready — Excel output: %s", _EXCEL_LOG.resolve())
        except PermissionError:
            print(
                f"WARNING: {_EXCEL_LOG} is locked (Excel has it open). "
                "Close Excel before running the bot, or Excel updates will be skipped.",
                flush=True,
            )
            log.warning(
                "Excel file locked (%s). Close Excel before running the bot.", _EXCEL_LOG
            )

    # ── Public API ────────────────────────────────────────────────────────────

    def log_trade(
        self,
        *,
        symbol:           str,
        direction:        str,        # "LONG" | "SHORT"
        entry_time:       datetime,
        entry_price:      float,
        exit_time:        datetime,
        exit_price:       float,
        quantity:         int,
        strategy_name:    str,
        confluence_score: int,
        stop_loss:        float,
        take_profit:      float,
        exit_reason:      str,        # "TP1"|"TP2"|"TP3"|"SL"|"EOD"|"MANUAL"
        confluences_used: Optional[str] = None,   # comma-separated confluence names
        bias_score:       Optional[int] = None,
        kill_zone:        bool          = False,
    ) -> None:
        mult      = 1 if direction == "LONG" else -1
        gross_pnl = round((exit_price - entry_price) * mult * quantity, 2)

        row = [
            entry_time.strftime("%Y-%m-%d"),
            symbol,
            direction,
            entry_time.strftime("%Y-%m-%d %H:%M:%S"),
            round(entry_price, 4),
            exit_time.strftime("%Y-%m-%d %H:%M:%S"),
            round(exit_price,  4),
            quantity,
            gross_pnl,
            strategy_name,
            confluence_score,
            round(stop_loss,   4),
            round(take_profit, 4),
            exit_reason,
            confluences_used or "",
            bias_score if bias_score is not None else "",
            "Yes" if kill_zone else "No",
        ]

        with self._lock:
            with open(_TRADE_LOG, "a", newline="") as fh:
                csv.writer(fh).writerow(row)
            trades = self._read_all_trades()
            self._rebuild_daily_summary(trades)
            self._rebuild_excel(trades)

        log.info(
            "[TradeLog] %s %s x%d  entry=%.4f  exit=%.4f  pnl=%.2f  reason=%s",
            symbol, direction, quantity, entry_price, exit_price, gross_pnl, exit_reason,
        )

    # ── Internal helpers ──────────────────────────────────────────────────────

    def _read_all_trades(self) -> list[dict]:
        trades: list[dict] = []
        try:
            with open(_TRADE_LOG, newline="") as fh:
                for row in csv.DictReader(fh):
                    try:
                        row["gross_pnl"] = float(row["gross_pnl"])
                    except (ValueError, KeyError):
                        row["gross_pnl"] = 0.0
                    trades.append(row)
        except Exception as exc:
            log.warning("Could not read trade log: %s", exc)
        return trades

    def _rebuild_daily_summary(self, trades: list[dict]) -> None:
        """Re-aggregate by date and overwrite daily_summary.csv. Lock must be held."""
        daily = _aggregate_daily(trades)

        rows = []
        for d in sorted(daily):
            s     = daily[d]
            total = s["total_trades"]
            rows.append([
                d, total, s["winners"], s["losers"],
                round(s["gross_pnl"], 2),
                round(s["winners"] / total * 100, 1) if total else 0.0,
                round(s["best_trade"]  or 0.0, 2),
                round(s["worst_trade"] or 0.0, 2),
            ])

        with open(_DAILY_LOG, "w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(_DAILY_COLS)
            w.writerows(rows)

    def _rebuild_excel(self, trades: list[dict]) -> None:
        """Build APEX_Trade_Log.xlsx with 3 sheets. Lock must be held."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            log.warning("openpyxl not installed; skipping Excel export. Run: pip install openpyxl")
            return

        GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        wb = openpyxl.Workbook()

        # ── Sheet 1: Trade History ────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Trade History"
        h1 = [
            "Date", "Symbol", "Direction", "Entry Time", "Entry Price",
            "Exit Time", "Exit Price", "Quantity", "Gross P&L ($)", "Result",
            "Exit Reason", "Strategy Used", "Confluences Used", "Bias Score", "Kill Zone",
        ]
        _write_header(ws1, h1)
        ws1.freeze_panes = "A2"
        ws1.auto_filter.ref = f"A1:{get_column_letter(len(h1))}1"

        for t in trades:
            pnl    = t["gross_pnl"]
            result = "W" if pnl > 0 else ("L" if pnl < 0 else "B")
            ws1.append([
                t.get("date", ""),
                t.get("symbol", ""),
                t.get("direction", ""),
                t.get("entry_time", ""),
                _to_float(t.get("entry_price")),
                t.get("exit_time", ""),
                _to_float(t.get("exit_price")),
                _to_int(t.get("quantity")),
                pnl,
                result,
                t.get("exit_reason", ""),
                t.get("strategy_name", ""),
                t.get("confluences_used", ""),
                _to_int(t.get("bias_score")) if t.get("bias_score") else "",
                t.get("kill_zone", ""),
            ])
            fill = GREEN if pnl > 0 else (RED if pnl < 0 else None)
            if fill:
                for cell in ws1[ws1.max_row]:
                    cell.fill = fill

        _autofit(ws1)

        # ── Sheet 2: Daily Summary ────────────────────────────────────────────
        ws2 = wb.create_sheet("Daily Summary")
        h2 = [
            "Date", "Total Trades", "Winners", "Losers", "Win Rate %",
            "Gross P&L", "Best Trade ($)", "Worst Trade ($)",
        ]
        _write_header(ws2, h2)
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(h2))}1"

        daily = _aggregate_daily(trades)
        for d in sorted(daily):
            s     = daily[d]
            total = s["total_trades"]
            ws2.append([
                d,
                total,
                s["winners"],
                s["losers"],
                round(s["winners"] / total * 100, 1) if total else 0.0,
                round(s["gross_pnl"], 2),
                round(s["best_trade"]  or 0.0, 2),
                round(s["worst_trade"] or 0.0, 2),
            ])

        _autofit(ws2)

        # ── Sheet 3: Weekly Summary ───────────────────────────────────────────
        ws3 = wb.create_sheet("Weekly Summary")
        h3 = [
            "Week", "Total Trades", "Win Rate %", "Total P&L",
            "Best Day ($)", "Worst Day ($)",
        ]
        _write_header(ws3, h3)
        ws3.freeze_panes = "A2"
        ws3.auto_filter.ref = f"A1:{get_column_letter(len(h3))}1"

        weekly = _aggregate_weekly(daily)
        for wk in sorted(weekly):
            s     = weekly[wk]
            total = s["total_trades"]
            ws3.append([
                wk,
                total,
                round(s["winners"] / total * 100, 1) if total else 0.0,
                round(s["gross_pnl"], 2),
                round(s["best_day"]  or 0.0, 2),
                round(s["worst_day"] or 0.0, 2),
            ])

        _autofit(ws3)

        try:
            wb.save(_EXCEL_LOG)
            log.info("Excel workbook updated: %s", _EXCEL_LOG)
        except PermissionError:
            log.warning(
                "Excel file locked by another process (%s) — "
                "trade recorded in CSV, Excel skipped this cycle. Close Excel to re-enable.",
                _EXCEL_LOG,
            )


# ── Module-level helpers ──────────────────────────────────────────────────────

def _write_header(ws, headers: list[str]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    ws.append(headers)
    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = align


def _autofit(ws) -> None:
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def _to_float(val) -> Optional[float]:
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _to_int(val) -> Optional[int]:
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


def _aggregate_daily(trades: list[dict]) -> dict:
    daily: dict = defaultdict(lambda: {
        "total_trades": 0, "winners": 0, "losers": 0,
        "gross_pnl": 0.0, "best_trade": None, "worst_trade": None,
    })
    for t in trades:
        d   = t.get("date", "")
        pnl = t["gross_pnl"]
        daily[d]["total_trades"] += 1
        daily[d]["gross_pnl"]    += pnl
        if pnl > 0:
            daily[d]["winners"] += 1
        elif pnl < 0:
            daily[d]["losers"]  += 1
        b = daily[d]["best_trade"]
        w = daily[d]["worst_trade"]
        daily[d]["best_trade"]  = pnl if b is None else max(pnl, b)
        daily[d]["worst_trade"] = pnl if w is None else min(pnl, w)
    return daily


def _aggregate_weekly(daily: dict) -> dict:
    weekly: dict = defaultdict(lambda: {
        "total_trades": 0, "winners": 0, "losers": 0,
        "gross_pnl": 0.0, "best_day": None, "worst_day": None,
    })
    for d, s in daily.items():
        try:
            dt = datetime.strptime(d, "%Y-%m-%d")
            wk = f"{dt.year}-W{dt.strftime('%W').zfill(2)}"
        except ValueError:
            continue
        day_pnl = s["gross_pnl"]
        weekly[wk]["total_trades"] += s["total_trades"]
        weekly[wk]["winners"]      += s["winners"]
        weekly[wk]["losers"]       += s["losers"]
        weekly[wk]["gross_pnl"]    += day_pnl
        b = weekly[wk]["best_day"]
        w = weekly[wk]["worst_day"]
        weekly[wk]["best_day"]  = day_pnl if b is None else max(day_pnl, b)
        weekly[wk]["worst_day"] = day_pnl if w is None else min(day_pnl, w)
    return weekly


# Module-level singleton — import this everywhere
_logger = TradeLogger()
